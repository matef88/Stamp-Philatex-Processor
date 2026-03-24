#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Photo Inventory Scanner
Scans a folder for images and creates/updates an Excel inventory with Arabic headers,
RTL support, and clickable hyperlinks to the files.
"""

import os
import sys
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


def setup_excel_headers(ws):
    """Set up the Excel headers with Arabic text and RTL formatting"""
    headers = [
        "#",  # A - Serial Number
        "Type",  # B - Type
        "الدولة",  # C - Country
        "الثيم",  # D - Theme
        "الوصف",  # E - Description
        "سنة الاصدار",  # F - Release Year
        "الحالة",  # G - Condition
        "عدد",  # H - Count
        "MiNr",  # I - Michel Number
        "SG",  # J - Stanley Gibbons
        "قيمة الكتالوج",  # K - Catalog Value
        "قيمة الكتالوج بالريال",  # L - Catalog Value in Riyal
        "النسبة",  # M - Percentage
        "سعر بيع الكتالوج",  # N - Catalog Selling Price
        "سعر البيع",  # O - Selling Price
        "اسم الفايل في الفولدر",  # P - Filename in Folder
        "هايبر لينك"  # Q - Hyperlink
    ]

    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, size=11)
        cell.alignment = Alignment(horizontal='right', vertical='center')

    # Set RTL direction for the sheet
    ws.sheet_view.rightToLeft = True

    # Add dropdown validation for Theme column (D)
    theme_list = [
        "شخصيات",
        "صليب هلال احمر",
        "طيور",
        "أعياد الميلاد",
        "فنون",
        "نباتات",
        "UPU",
        "اولمبيات",
        "الحياة البحرية",
        "الطيور",
        "الحيوانات",
        "الفضاء",
        "فنون",
        "مناسبات دينية",
        "وسائل النقل",
        "الزهور",
        "فلسطين",
        "اصدار مشترك"
    ]

    # Create data validation for Theme column
    dv = DataValidation(type="list", formula1=f'"{",".join(theme_list)}"', allow_blank=True)
    dv.error = 'يرجى اختيار قيمة من القائمة'
    dv.errorTitle = 'قيمة غير صحيحة'
    dv.prompt = 'اختر من القائمة'
    dv.promptTitle = 'الثيم'

    # Apply validation to column D (starting from row 2 to row 10000)
    ws.add_data_validation(dv)
    dv.add(f'D2:D10000')

    return headers


def get_existing_filenames(ws):
    """Get list of existing filenames from column P to avoid duplicates"""
    existing_files = set()

    # Column P is column 16
    for row in range(2, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=16).value
        if cell_value:
            existing_files.add(str(cell_value).strip())

    return existing_files


def update_serial_numbers(ws):
    """Update serial numbers in column A for all rows"""
    for row in range(2, ws.max_row + 1):
        serial_number = row - 1  # Row 2 gets #1, row 3 gets #2, etc.
        ws.cell(row=row, column=1).value = serial_number
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')


def auto_fit_columns(ws):
    """Auto-fit all columns based on content"""
    for column_cells in ws.columns:
        length = 0
        column = column_cells[0].column_letter
        
        for cell in column_cells:
            try:
                if cell.value:
                    # Calculate length considering Arabic text
                    cell_value = str(cell.value)
                    length = max(length, len(cell_value))
            except:
                pass
        
        # Set column width with minimum and maximum bounds
        adjusted_width = min(max(length + 2, 10), 50)
        ws.column_dimensions[column].width = adjusted_width


def scan_folder_for_images(folder_path):
    """Scan folder for image files"""
    valid_extensions = {'.jpg', '.jpeg', '.png', '.gif', '.bmp', '.webp'}
    image_files = []
    
    folder = Path(folder_path)
    
    if not folder.exists():
        print(f"❌ Error: Folder does not exist: {folder_path}")
        return []
    
    if not folder.is_dir():
        print(f"❌ Error: Path is not a directory: {folder_path}")
        return []
    
    # Scan for image files
    for file_path in folder.iterdir():
        if file_path.is_file() and file_path.suffix.lower() in valid_extensions:
            image_files.append(file_path)
    
    return sorted(image_files, key=lambda x: x.name.lower())


def add_file_to_excel(ws, file_path, row_num):
    """Add a file entry to the Excel with hyperlink and formulas"""
    filename = file_path.name
    absolute_path = str(file_path.absolute())

    # Column A (1): Serial Number
    serial_number = row_num - 1  # Row 2 gets serial #1, row 3 gets #2, etc.
    ws.cell(row=row_num, column=1).value = serial_number
    ws.cell(row=row_num, column=1).alignment = Alignment(horizontal='center')

    # Column L (12): Formula - Catalog Value in Riyal = K * 5
    ws.cell(row=row_num, column=12).value = f"=K{row_num}*5"
    ws.cell(row=row_num, column=12).alignment = Alignment(horizontal='right')

    # Column N (14): Formula - Catalog Selling Price = M * L
    ws.cell(row=row_num, column=14).value = f"=M{row_num}*L{row_num}"
    ws.cell(row=row_num, column=14).alignment = Alignment(horizontal='right')

    # Column P (16): Filename
    ws.cell(row=row_num, column=16).value = filename
    ws.cell(row=row_num, column=16).alignment = Alignment(horizontal='right')

    # Column Q (17): Hyperlink with "Open" text
    cell = ws.cell(row=row_num, column=17)
    cell.value = "Open"
    cell.hyperlink = absolute_path
    cell.font = Font(color="0563C1", underline="single")
    cell.alignment = Alignment(horizontal='center')

    # Set alignment for empty cells to RTL
    for col in range(1, 18):
        if col not in [1, 12, 14, 16, 17]:  # Skip the cells we already filled
            ws.cell(row=row_num, column=col).alignment = Alignment(horizontal='right')


def main():
    """Main execution function"""
    print("=" * 60)
    print("📸 Photo Inventory Scanner")
    print("=" * 60)
    print()
    
    # Get folder path from user
    folder_path = input("Enter the folder path to scan for images: ").strip()
    
    # Remove quotes if user pasted path with quotes
    folder_path = folder_path.strip('"').strip("'")
    
    if not folder_path:
        print("❌ Error: No folder path provided")
        sys.exit(1)
    
    print(f"\n🔍 Scanning folder: {folder_path}")
    
    # Scan for images
    image_files = scan_folder_for_images(folder_path)
    
    if not image_files:
        print("⚠️  No image files found in the specified folder")
        sys.exit(0)
    
    print(f"✅ Found {len(image_files)} image file(s)")

    # Excel file path (in the scanned folder with priority filename)
    excel_file = Path(folder_path) / "001 - Photo_Inventory.xlsx"

    # Load or create workbook
    if excel_file.exists():
        print(f"\n📂 Loading existing Excel file: {excel_file}")
        wb = load_workbook(excel_file)
        ws = wb.active
        existing_files = get_existing_filenames(ws)
        print(f"📋 Found {len(existing_files)} existing entries")
        # Update serial numbers for existing rows
        print(f"🔢 Updating serial numbers...")
        update_serial_numbers(ws)
    else:
        print(f"\n📝 Creating new Excel file: {excel_file}")
        wb = Workbook()
        ws = wb.active
        ws.title = "Photo Inventory"
        setup_excel_headers(ws)
        existing_files = set()
    
    # Add new files
    new_files_added = 0
    next_row = ws.max_row + 1
    
    print("\n🔄 Processing files...")
    
    for file_path in image_files:
        filename = file_path.name
        
        if filename in existing_files:
            print(f"⏭️  Skipping duplicate: {filename}")
            continue
        
        add_file_to_excel(ws, file_path, next_row)
        print(f"✅ Added: {filename}")
        new_files_added += 1
        next_row += 1
    
    # Auto-fit columns and save
    if new_files_added > 0:
        print("\n🔢 Updating all serial numbers...")
        update_serial_numbers(ws)

        print("📐 Auto-fitting columns...")
        auto_fit_columns(ws)

        # Save workbook
        print(f"💾 Saving Excel file...")
        wb.save(excel_file)
        print(f"✅ Successfully saved: {excel_file.absolute()}")
    else:
        print("\n⚠️  No new files to add (all files already exist in Excel)")
        # Still save to update serial numbers if they were updated
        if excel_file.exists():
            print(f"💾 Saving updated serial numbers...")
            wb.save(excel_file)
    
    # Summary
    print("\n" + "=" * 60)
    print("📊 Summary:")
    print(f"   Total images found: {len(image_files)}")
    print(f"   Already in Excel: {len(existing_files)}")
    print(f"   New files added: {new_files_added}")
    print("=" * 60)
    print()


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\n⚠️  Operation cancelled by user")
        sys.exit(0)
    except Exception as e:
        print(f"\n❌ Error: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
